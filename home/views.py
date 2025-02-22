from django.shortcuts import render,redirect,get_object_or_404
from .models import Product,ShoppingList
from django.http import HttpResponse
from django_daraja.mpesa.core import MpesaClient
from django.conf import settings
import uuid
import openpyxl
from openpyxl.styles import Font
from django.db.models import Q
import requests
import json
from django.http import JsonResponse
from openpyxl.utils import get_column_letter

mpesa_client = MpesaClient()

def mpesa_payment(request):
        if request.method == "POST":
            phone_number = request.POST.get("phone_number")
        if not phone_number:
            return HttpResponse("Phone number is required.")

        if not phone_number.startswith("254") or not phone_number.isdigit() or len(phone_number) != 12:
            return HttpResponse("Invalid phone number. Please use the format 2547XXXXXXXX.")
        amount = int(sum(item.quantity * item.product.Price for item in ShoppingList.objects.all()))
        account_reference = str(uuid.uuid4())[:8]
        transaction_desc = "Shopping List Payment"
        callback_url = 'https://darajambili.herokuapp.com/express-payment';

        response = mpesa_client.stk_push(
            phone_number, amount, account_reference, transaction_desc,callback_url
        )

        try:
            response_data = response.json()
            print("Response Data:", response_data)

            if response_data.get("ResponseCode") == "0":
                request.session["shopping_list_id"] = account_reference
                return redirect("home:download_shopping_list_after_payment")
            else:
                error_message = response_data.get("errorMessage", "Transaction failed.")
                return HttpResponse(f"Payment failed: {error_message}")
        except ValueError:
            return HttpResponse("An error occurred: Invalid response format.")

        return HttpResponse("Invalid request method.")

def generate_shopping_list_xlsx():
    """Generates an XLSX file of the shopping list."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Shopping List"
    
    # Define headers
    headers = ["Product", "Quantity", "Price (KSH)", "Total Price (KSH)"]
    sheet.append(headers)
    
    # Fetch shopping items
    shopping_items = ShoppingList.objects.all()
    total_price = 0
    
    for item in shopping_items:
        total_item_price = item.quantity * item.product.Price
        total_price += total_item_price
        sheet.append([
            item.product.name,
            item.quantity,
            item.product.Price,
            total_item_price
        ])
    
    # Add overall total row
    sheet.append(["", "", "Overall Total (KSH):", total_price])
    
    # Auto-adjust column width
    for col_num, column_cells in enumerate(sheet.columns, 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column].width = max_length + 2
    
    return workbook

def download_shopping_list_after_payment(request):
    """View to generate and return the shopping list as an XLSX file."""
    workbook = generate_shopping_list_xlsx()
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=shopping_list.xlsx'
    
    workbook.save(response)
    return response

def home(request):
    products = Product.objects.filter(quantity__gt=0)
    return render(request, 'home/homes.html', {
        'products':products,
    })

def shopping_list_view(request):
    shopping_items = ShoppingList.objects.all()

    for item in shopping_items:
        if not item.unique_id:
            item.unique_id = uuid.uuid4()
            item.save()

    total_price = sum(item.quantity * item.product.Price for item in shopping_items)

    # Add total cost for each item
    for item in shopping_items:
        item.total_cost = item.quantity * item.product.Price  # Compute in view

    context = {
        'shopping_items': shopping_items,
        'total_price': total_price,
    }

    return render(request, 'home/shopping_list.html', context)


def add_to_shopping_list(request, product_id):
    if request.method == "POST":
        product = get_object_or_404(Product, id=product_id)
        quantity = int(request.POST.get("quantity", 1))  
        
        shopping_item, created = ShoppingList.objects.get_or_create(
            product=product,
            defaults={'quantity': quantity}
        )
        if not created:
            shopping_item.quantity += quantity  
            shopping_item.save()

        return redirect('home:shopping_list')



def update_shopping_list(request, item_id):
    if request.method == "POST":
        shopping_item = get_object_or_404(ShoppingList, id=item_id)
        quantity = int(request.POST.get("quantity", 1))  
        shopping_item.quantity = quantity
        shopping_item.save()
        return redirect('home:shopping_list')

def delete_from_shopping_list(request, item_id):
    if request.method == "POST":
        shopping_item = get_object_or_404(ShoppingList, id=item_id)
        shopping_item.delete()
        return redirect('home:shopping_list')

def shopping_list_detail(request, unique_id):
    shopping_list = get_object_or_404(ShoppingList, unique_id=unique_id)
    items = shopping_list.items.all()

    context = {
        'shopping_list': shopping_list,
        'items': items,
        'total_price': shopping_list.total_price(),
    }
    return render(request, 'home/shopping_list_detail.html', context)

def detail(request,pk):
    item = get_object_or_404(Product, pk=pk)
    return render(request, 'home/detail.html', {
        'item':item,
    })


def search(request):
    query = request.GET.get('query','')
    items = []
    if query:
        query_words = query.split()
        items_query = Q()
        for word in query_words:
            items_query &= Q(name__icontains=word) | Q(description__icontains=word)
        items = Product.objects.filter(items_query)
    return render(request,'home/search.html',{
        'query':query,
        'items':items,
    }) 
    
    
def help(request):
    return render(request, 'home/help.html') 

def initiate_payment(request, unique_id):
    shopping_items = ShoppingList.objects.all()

    total_price = sum(item.quantity * item.product.Price for item in shopping_items)

    # Add total cost for each item
    for item in shopping_items:
        item.total_cost = item.quantity * item.product.Price  # Compute in view

    # Paystack details
    paystack_secret_key = settings.PAYSTACK_SECRET_KEY
    callback_url = request.build_absolute_uri('/')

    payment_data = {
        "email": f"user_{uuid.uuid4()}@example.com",  # Temporary email
        "amount": int(total_price * 100),  # Convert to kobo
        "currency": "KES",
        "callback_url": callback_url
    }

    headers = {
        "Authorization": f"Bearer {paystack_secret_key}",
        "Content-Type": "application/json"
    }

    response = requests.post(
        "https://api.paystack.co/transaction/initialize",
        headers=headers,
        data=json.dumps(payment_data)
    )
    
    if response.status_code == 200:
        
        response_data = response.json()
        return redirect(response_data['data']['authorization_url'])  # Redirect to Paystack

    return JsonResponse({"error": "Payment initiation failed"}, status=400)
    
def payment_success(request):
    return redirect('home:download_shopping_list_after_payment')